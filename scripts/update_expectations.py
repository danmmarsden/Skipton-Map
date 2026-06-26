from __future__ import annotations

import json
import math
import pathlib
import re
import urllib.error
import urllib.request

import openpyxl


ROOT = pathlib.Path(__file__).resolve().parents[1]
HTML_PATH = ROOT / "index.html"
DATA_PATH = ROOT / "map-data.js"
WORKBOOK_PATH = ROOT / "Skipton - CGI partners travel.xlsx"
SHEET_NAME = "Final View"

EARTH_RADIUS_MILES = 3958.7613
FULL_POSTCODE_CACHE = {
    "LS25 6PQ": [53.795993, -1.23728],
    "YO42 2LY": [53.932113, -0.773118],
    "PR1 9HS": [53.739046, -2.72244],
    "HA2 9QL": [51.569997, -0.381045],
    "M33 7GG": [53.424346, -2.322915],
    "LS16 6JN": [53.840085, -1.598052],
}


def clean_text(value):
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


def clean_postcode(value):
    text = clean_text(value)
    if not text or text.upper() in {"N/A", "NA", "NONE"}:
        return None
    return re.sub(r"\s+", " ", text.upper())


def outward_code(postcode):
    if not postcode:
        return None

    compact = re.sub(r"\s+", "", postcode.upper())
    match = re.match(r"^([A-Z]{1,2}\d[A-Z\d]?)(\d[A-Z]{2})$", compact)
    return match.group(1) if match else compact


def is_full_postcode(postcode):
    return bool(postcode and re.match(r"^[A-Z]{1,2}\d[A-Z\d]?\s?\d[A-Z]{2}$", postcode))


def map_locations():
    html = HTML_PATH.read_text()

    skipton_match = re.search(r"const skipton = \[([^\]]+)\];", html)
    if not skipton_match:
        raise RuntimeError("Could not find SBS coordinates in index.html")

    skipton = [float(value.strip()) for value in skipton_match.group(1).split(",")]

    data_text = DATA_PATH.read_text()
    centroids_match = re.search(
        r'"districtCentroids": (.*?),\n  "fullPostcodeLocations"',
        data_text,
        flags=re.S,
    )
    if not centroids_match:
        raise RuntimeError("Could not find district centroids in map-data.js")

    return skipton, json.loads(centroids_match.group(1))


def fetch_full_postcode(postcode):
    if postcode in FULL_POSTCODE_CACHE:
        return FULL_POSTCODE_CACHE[postcode]

    url = "https://api.postcodes.io/postcodes/" + urllib.request.quote(postcode)
    try:
        with urllib.request.urlopen(url, timeout=5) as response:
            body = json.loads(response.read().decode("utf-8"))
    except (urllib.error.URLError, TimeoutError):
        return None

    result = body.get("result")
    if not result:
        return None

    return [result["latitude"], result["longitude"]]


def postcode_location(postcode, centroids):
    if not postcode:
        return None

    if is_full_postcode(postcode):
        coords = fetch_full_postcode(postcode)
        if coords:
            return coords

    return centroids.get(outward_code(postcode))


def miles_between(a, b):
    lat1, lon1 = [math.radians(value) for value in a]
    lat2, lon2 = [math.radians(value) for value in b]
    dlat = lat2 - lat1
    dlon = lon2 - lon1

    h = (
        math.sin(dlat / 2) ** 2
        + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    )
    return 2 * EARTH_RADIUS_MILES * math.asin(math.sqrt(h))


def expectation_for(distance):
    if distance <= 30:
        return "3 days per week"
    if distance <= 40:
        return "2 days per week"
    return "1 day per week"


def main():
    skipton, centroids = map_locations()
    workbook = openpyxl.load_workbook(WORKBOOK_PATH)
    sheet = workbook[SHEET_NAME]
    headers = [cell.value for cell in sheet[1]]
    columns = {header: index + 1 for index, header in enumerate(headers)}

    postcode_col = columns["Postcode"]
    expectation_col = columns["Expectation based on location"]
    name_col = columns["Name"]

    counts = {"3 days per week": 0, "2 days per week": 0, "1 day per week": 0}
    skipped = []
    updated = []

    for row_number in range(2, sheet.max_row + 1):
        name = clean_text(sheet.cell(row_number, name_col).value)
        if not name:
            continue

        postcode = clean_postcode(sheet.cell(row_number, postcode_col).value)
        coords = postcode_location(postcode, centroids)

        if not coords:
            sheet.cell(row_number, expectation_col).value = None
            skipped.append({"row": row_number, "name": name, "postcode": postcode})
            continue

        distance = miles_between(skipton, coords)
        expectation = expectation_for(distance)
        sheet.cell(row_number, expectation_col).value = expectation
        counts[expectation] += 1
        updated.append(
            {
                "row": row_number,
                "name": name,
                "postcode": postcode,
                "distance": round(distance, 2),
                "expectation": expectation,
            }
        )

    workbook.save(WORKBOOK_PATH)

    print(json.dumps({"updated": len(updated), "skipped": skipped, "counts": counts}, indent=2))


if __name__ == "__main__":
    main()
